import os
import nibabel as nib
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import math
from collections import Counter
from openpyxl.styles import Font, colors, Alignment
import cv2

Train_path = 'MICCAI_BraTS_2019_Data_Training/seg/HGG'
Train_orig_path = 'MICCAI_BraTS_2019_Data_Training/HGG'

Valid_path = 'MICCAI_BraTS_2019_Data_Validation/seg'
Valid_orig_path = 'MICCAI_BraTS_2019_Data_Validation'

Test_path = 'MICCAI_BraTS_2019_Data_Testing_20190909/Test'
Test_orig_path = 'MICCAI_BraTS_2019_Data_Testing_20190909/MICCAI_BraTS_2019_Data_Testing'

Train_reg = 'MICCAI_BraTS_2019_Data_Training_6_26_27_57Label/HGG'
Valid_reg = 'MICCAI_BraTS_2019_Data_Validation_27_57'
Test_reg = 'MICCAI_BraTS_2019_Data_Testing_20190909/MICCAI_BraTS_2019_Data_Testing'

wb = load_workbook('aa.xlsx')
ws = wb.active
train_dir = []
for i in ws['A']:
    train_dir.append(i.value)
train_dir = train_dir[1:]

train_age = []
for i in ws['B']:
    train_age.append(i.value)
train_age = train_age[1:]

train_RS = []
for i in ws['D']:
    train_RS.append(i.value)
train_RS = train_RS[1:]

print(len(train_dir))

# seg label 1--->necrotic
#           2--->edema
#           4--->enhance

def count_volume(path, orig_path, path_reg):
    dir = sorted(os.listdir(path))
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 25.0
    title = ['ID', 'Whole', 'Necrotic', 'Edema', 'Enhance', 'WholeBrain', 'NecroticBrain', 'EdemaBrain', 'EnhanceBrain', 'NecroticEnhance', 'EdemaEnhance', 'NecroticEdema',
             'Whole_grad', 'Necrotic_grad', 'Edema_grad', 'Enhance_grad', 'Whole_SVRatio', 'Necrotic_SVRatio', 'Edema_SVRatio', 'Enhance_SVRatio',
             'Whole_ID', 'True_ID', 'Whole_X', 'Whole_Y', 'Whole_Z', 'True_X', 'True_Y', 'True_Z', 
             'Brain_tumor_core_X', 'Brain_tumor_core_Y', 'Brain_tumor_core_Z', 'Brain_enh_core_X', 'Brain_enh_core_Y', 'Brain_enh_core_Z', 'Age', 'RS0', 'RS1']
    ws.append(title)
    for (i, j, k) in zip(train_dir, train_age, train_RS):
        print(k)
        write = []
        name = i.split('.nii.gz')[0]
        image = nib.load(os.path.join(path, name+'.nii.gz')).get_data()
        region = nib.load(os.path.join(path_reg, name, name+'_27labelsJointFusion.nii.gz')).get_data()
        orig_image = nib.load(os.path.join(orig_path, name, name+'_t1ce.nii.gz')).get_data()
        print(region.shape)

        data = (np.array([1,2,4]) == np.array(image)[...,None]).astype(int)
        whole = image.copy()
        whole[whole>0] = 1
        whole_brain = orig_image.copy()
        whole_brain[whole_brain>0] = 1
        nec_volume = np.sum(data[...,0])
        ede_volume = np.sum(data[...,1])
        enh_volume = np.sum(data[...,2])

        Whole_Brain = np.sum(whole)/np.sum(whole_brain)
        nec_brain = nec_volume/np.sum(whole_brain)
        ede_brain = ede_volume/np.sum(whole_brain)
        enh_brain = enh_volume/np.sum(whole_brain)

        nec_enh_volume = nec_volume/enh_volume
        ede_enh_volume = ede_volume/enh_volume
        nec_ede_volume = nec_volume/ede_volume

        necrotic = data[..., 0]
        edema = data[..., 1]
        enhance = data[..., 2]
        
        grad_image = np.zeros(image.shape)
        for layer in range(grad_image.shape[-1]):
            sobelx = cv2.Sobel(np.float32(np.squeeze(necrotic[:,:,layer])),cv2.CV_64F, 1, 0, ksize=3)
            sobely = cv2.Sobel(np.float32(np.squeeze(necrotic[:,:,layer])),cv2.CV_64F, 0, 1,  ksize=3)
            grad_image[:,:,layer] = np.sqrt(sobelx**2+sobely**2)

            sobelx = cv2.Sobel(np.float32(np.squeeze(edema[:,:,layer])),cv2.CV_64F, 1, 0, ksize=3)
            sobely = cv2.Sobel(np.float32(np.squeeze(edema[:,:,layer])),cv2.CV_64F, 0, 1,  ksize=3)
            grad_image[:,:,layer] += np.sqrt(sobelx**2+sobely**2)

            sobelx = cv2.Sobel(np.float32(np.squeeze(enhance[:,:,layer])),cv2.CV_64F, 1, 0, ksize=3)
            sobely = cv2.Sobel(np.float32(np.squeeze(enhance[:,:,layer])),cv2.CV_64F, 0, 1,  ksize=3)
            grad_image[:,:,layer] += np.sqrt(sobelx**2+sobely**2)    
        grad_image[grad_image>0] = 1
        grad_image = grad_image * image
        grad_image = (np.array([1,2,4]) == np.array(grad_image)[...,None]).astype(int)
        whole_grad_image = grad_image.copy()
        whole_grad_image[whole_grad_image>0] = 1
        nec_grad_volume = np.sum(grad_image[...,0])
        ede_grad_volume = np.sum(grad_image[...,1])
        enh_grad_volume = np.sum(grad_image[...,2])
        
        nec_SVRatio = nec_grad_volume/nec_volume
        ede_SVRatio = ede_grad_volume/ede_volume
        enh_SVRatio = enh_grad_volume/enh_volume

        whole_volume = np.sum(whole)
        whole_coord = np.mean(np.where(whole == 1), 1)
        whole_brain_coord = np.mean(np.where(whole_brain == 1), 1)

        whole_grad_volume = np.sum(whole_grad_image)
        whole_SVRatio = whole_grad_volume/whole_volume

        X = [math.floor(whole_coord[0]), math.ceil(whole_coord[0])]
        Y = [math.floor(whole_coord[1]), math.ceil(whole_coord[1])]
        Z = [math.floor(whole_coord[2]), math.ceil(whole_coord[2])]

        Region_value = [region[i][j][k] for i in X for j in Y for k in Z]
        ID_value = Counter(Region_value)
        whole_value = ID_value.most_common(1)[0][0]
        

        if enh_volume == 0:
            true_coord = whole_coord
        else:
            true_coord = np.mean(np.where(enhance == 1), 1)

        True_X = [math.floor(true_coord[0]), math.ceil(true_coord[0])]
        True_Y = [math.floor(true_coord[1]), math.ceil(true_coord[1])]
        True_Z = [math.floor(true_coord[2]), math.ceil(true_coord[2])]

        Region_value = [region[i][j][k] for i in True_X for j in True_Y for k in True_Z]
        ID_value = Counter(Region_value)
        true_value = ID_value.most_common(1)[0][0]

        print(whole_value,true_value)

        write.append(name)
        write.append(whole_volume)
        write.append(nec_volume)
        write.append(ede_volume)
        write.append(enh_volume)

        write.append(Whole_Brain)
        write.append(nec_brain)
        write.append(ede_brain)
        write.append(enh_brain)
        
        write.append(nec_enh_volume)
        write.append(ede_enh_volume)
        write.append(nec_ede_volume)

        write.append(whole_grad_volume)
        write.append(nec_grad_volume)
        write.append(ede_grad_volume)
        write.append(enh_grad_volume)

        write.append(whole_SVRatio)
        write.append(nec_SVRatio)
        write.append(ede_SVRatio)
        write.append(enh_SVRatio)

        write.append(whole_value)
        write.append(true_value)

        write.append(round(whole_coord[0], 1))
        write.append(round(whole_coord[1], 1))
        write.append(round(whole_coord[2], 1))

        write.append(round(true_coord[0], 1))
        write.append(round(true_coord[1], 1))
        write.append(round(true_coord[2], 1))

        write.append(round(whole_coord[0], 1) - round(whole_brain_coord[0], 1))
        write.append(round(whole_coord[1], 1) - round(whole_brain_coord[1], 1))
        write.append(round(whole_coord[2], 1) - round(whole_brain_coord[2], 1))

        write.append(round(true_coord[0], 1) - round(whole_brain_coord[0], 1))
        write.append(round(true_coord[1], 1) - round(whole_brain_coord[1], 1))
        write.append(round(true_coord[2], 1) - round(whole_brain_coord[2], 1))

        write.append(j)
        if k == 'GTR':
            write.append(int(1))
            write.append(int(0))
        elif k == 'STR':
            write.append(int(0))
            write.append(int(1))
        else:
            write.append(int(0))
            write.append(int(0))
            
        ws.append(write)

        print(i)

#    wb.save(filename="Train_final.xlsx")
#count_volume(Train_path, Train_orig_path, Train_reg)

#    wb.save(filename="Valid_final.xlsx")
#count_volume(Valid_path, Valid_orig_path, Valid_reg)

    wb.save(filename="Test_final.xlsx")
count_volume(Test_path, Test_orig_path, Test_reg)
